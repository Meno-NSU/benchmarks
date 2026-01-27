from pathlib import Path
import json
import openpyxl
import pandas as pd
from collections import defaultdict


def make_report(
    scan_dir: Path,
    out_file: Path = Path("summary"),
):
    scan_dir = Path(scan_dir).absolute()
    data = defaultdict(lambda: defaultdict(list))
    for parent in scan_dir.iterdir():
        for path in parent.glob("*_judged.json"):
            name = parent.name
            with path.open() as f:
                loaded = json.load(f)
            data_inner = data[name]
            for item in loaded:
                for k, v in item["case"].items():
                    data_inner[k].append(v)
                for k, v in item["result"].items():
                    if isinstance(v, dict):
                        for k2, v2 in v.items():
                            data_inner[f"{k}.{k2}"].append(str(v2))
    wb = openpyxl.Workbook()
    sh = wb.create_sheet("Report")
    sh = wb.active
    assert sh is not None
    if data:
        columns = ["model", *data[next(iter(data))].keys()]
        sh.append(columns)
        current_row = 2
        for parent, v in data.items():
            written_row_count = 0
            for row in zip(*v.values()):
                sh.append([parent, *row])
                written_row_count += 1
            sh.merge_cells(f"A{current_row}:A{current_row + written_row_count - 1}")
            current_row += written_row_count
    wb.save(out_file.with_suffix(".xlsx"))
    wb.close()
